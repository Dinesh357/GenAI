
from openpyxl import load_workbook
import os
from openpyxl import Workbook

def generate_output_excel(answers, scores, category):
    wb = load_workbook("data/Platform-Qualification-Checklst.xlsx")
    sheet = wb["Qualification-Checklist"]

    for row in sheet.iter_rows(min_row=13, max_row=31):
        criterion = row[0].value
        if criterion in answers:
            row[2].value = scores[criterion]
            try:
                weight = float(row[3].value)
            except:
                weight = 0
            row[4].value = weight * scores[criterion]
            row[5].value = answers[criterion]

    weighted_scores = []
    for row in sheet.iter_rows(min_row=13, max_row=31):
        try:
            weighted_scores.append(float(row[4].value))
        except:
            continue

    total_score = sum(weighted_scores)
    sheet["F33"] = total_score

    if total_score > 850:
        category_text = "Enterprise-grade Platform"
    elif total_score > 600:
        category_text = "Emerging Platform"
    elif total_score > 300:
        category_text = "Application / Modular Product"
    else:
        category_text = "Application Development"

    sheet["F36"] = f"Project classified as: {category_text}"

    output_path = "data/Platform_Qualification_Result.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


def generate_result_excel(rows, total_score, category, output_path: str = "data/Platform_Qualification_Result.xlsx") -> str:
    """
    Create a minimal results workbook from computed rows.
    Rows should contain: criterion, question, score, weight, weighted, justification
    Returns the saved file path.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    headers = ["Criterion", "Question", "Score", "Weight", "Weighted Score", "Justification"]
    ws.append(headers)
    for r in rows:
        ws.append([
            r.get("criterion"),
            r.get("question"),
            r.get("score"),
            r.get("weight"),
            r.get("weighted"),
            r.get("justification"),
        ])
    ws.append([])
    ws.append(["Total", None, None, None, total_score, None])
    ws.append(["Classification", category])
    wb.save(output_path)
    return output_path
