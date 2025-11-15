from openpyxl import load_workbook
from typing import List, Dict, Tuple
import os
import time


def load_checklist(path: str = "data/Platform-Qualification-Checklist.xlsx",
                   sheet_name: str = "Qualification-Checklist") -> List[Dict[str, object]]:
    """
    Load checklist items from the Excel template.
    Expects a header row containing: Criterion | Questions/Vocabulary | Score (S: 0-3) | Weight (W)
    Returns a list of dicts: {criterion, question, weight}
    """
    start = time.perf_counter()
    exists = os.path.exists(path)
    print(f"[Checklist] path={path} exists={exists}", flush=True)
    wb = load_workbook(path)
    print(f"[Checklist] sheetnames={wb.sheetnames} requested={sheet_name}", flush=True)
    if sheet_name not in wb.sheetnames:
        # fallback to first sheet
        sheet = wb[wb.sheetnames[0]]
        print(f"[Checklist] using_sheet={sheet.title} (fallback)", flush=True)
    else:
        sheet = wb[sheet_name]
        print(f"[Checklist] using_sheet={sheet.title}", flush=True)

    # Find header row (cell with value 'Criterion')
    header_row_idx = None
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        if (row and row[0].value and isinstance(row[0].value, str)
                and row[0].value.strip().lower() == "criterion"):
            header_row_idx = row[0].row
            break

    if header_row_idx is None:
        # Try default known header at row 12 (CSV shows header on row 12)
        header_row_idx = 12
    print(f"[Checklist] header_row_idx={header_row_idx}", flush=True)

    items: List[Dict[str, object]] = []
    # Columns: A=criterion, B=question, D=weight
    for row in sheet.iter_rows(min_row=header_row_idx + 1, max_row=sheet.max_row):
        criterion = (row[0].value or "").strip() if isinstance(row[0].value, str) else row[0].value
        question = (row[1].value or "").strip() if isinstance(row[1].value, str) else row[1].value
        weight_cell = row[3].value if len(row) >= 4 else None
        if not criterion:
            # Stop when we hit the blank area after the list
            continue
        try:
            weight = float(weight_cell) if weight_cell is not None and str(weight_cell).strip() != "" else 0.0
        except Exception:
            weight = 0.0

        items.append({
            "criterion": criterion,
            "question": question if question else str(criterion),
            "weight": weight,
        })

    elapsed = time.perf_counter() - start
    print(f"[Checklist] parsed_items={len(items)} elapsed={elapsed:.2f}s", flush=True)
    return items
